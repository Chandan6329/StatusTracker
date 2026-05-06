from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from datetime import datetime
from ..database import get_db
from ..models import Achievement, Task
from ..schemas import ExportRequest
import openpyxl
from openpyxl.styles import Font, PatternFill
import tempfile
import os

router = APIRouter()

@router.post("/export/achievements")
def export_achievements(request: ExportRequest, db: Session = Depends(get_db)):
    # Query achievements with optional date filtering
    query = db.query(Achievement)
    if request.start_date:
        query = query.filter(Achievement.timestamp >= request.start_date)
    if request.end_date:
        query = query.filter(Achievement.timestamp <= request.end_date)

    achievements = query.all()

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Achievements"

    # Header row
    headers = ["ID", "Description", "Category", "Timestamp", "Tags"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Data rows
    for row_num, achievement in enumerate(achievements, 2):
        ws.cell(row=row_num, column=1, value=achievement.id)
        ws.cell(row=row_num, column=2, value=achievement.description)
        ws.cell(row=row_num, column=3, value=achievement.category)
        ws.cell(row=row_num, column=4, value=achievement.timestamp.strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row_num, column=5, value=achievement.tags)

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    return FileResponse(
        tmp_path,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename=f'achievements_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

@router.post("/export/tasks")
def export_tasks(request: ExportRequest, db: Session = Depends(get_db)):
    # Query tasks
    query = db.query(Task)
    if request.start_date:
        query = query.filter(Task.created_date >= request.start_date)
    if request.end_date:
        query = query.filter(Task.created_date <= request.end_date)

    tasks = query.all()

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"

    # Header row
    headers = ["ID", "Description", "Priority", "Status", "Due Date", "Created Date"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

    # Data rows
    for row_num, task in enumerate(tasks, 2):
        ws.cell(row=row_num, column=1, value=task.id)
        ws.cell(row=row_num, column=2, value=task.description)
        ws.cell(row=row_num, column=3, value=task.priority.value)
        ws.cell(row=row_num, column=4, value=task.status.value)
        ws.cell(row=row_num, column=5, value=task.due_date.strftime("%Y-%m-%d") if task.due_date else "")
        ws.cell(row=row_num, column=6, value=task.created_date.strftime("%Y-%m-%d %H:%M:%S"))

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    return FileResponse(
        tmp_path,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        filename=f'tasks_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )