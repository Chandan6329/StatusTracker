from flask import Flask, request, jsonify, send_file, after_this_request
from .db import init_db, get_db, close_connection
from datetime import datetime
import tempfile
import os
import openpyxl
from openpyxl.styles import Font, PatternFill

app = Flask(__name__)
init_db()
app.teardown_appcontext(close_connection)

@app.after_request
def add_cors(response):
    response.headers["Access-Control-Allow-Origin"] = "http://localhost:3000"
    response.headers["Access-Control-Allow-Headers"] = "*"
    response.headers["Access-Control-Allow-Methods"] = "GET,POST,PUT,DELETE,OPTIONS"
    return response

@app.route("/health")
def health_check():
    return jsonify(status="healthy")

@app.route("/api/achievements", methods=["GET", "POST"])
def achievements():
    if request.method == "POST":
        data = request.get_json() or {}
        description = data.get("description", "").strip()
        category = data.get("category", "").strip()
        tags = data.get("tags", [])

        if not description or not category:
            return jsonify({"error": "description and category are required"}), 400

        timestamp = datetime.utcnow().isoformat()
        tags_csv = ",".join(tags) if isinstance(tags, list) else str(tags)
        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            "INSERT INTO achievements (description, category, timestamp, tags) VALUES (?, ?, ?, ?)",
            (description, category, timestamp, tags_csv)
        )
        db.commit()
        achievement_id = cursor.lastrowid
        cursor.execute("SELECT * FROM achievements WHERE id = ?", (achievement_id,))
        achievement = cursor.fetchone()
        return jsonify(dict(achievement)), 201

    skip = int(request.args.get("skip", 0))
    limit = int(request.args.get("limit", 100))
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT * FROM achievements ORDER BY timestamp DESC LIMIT ? OFFSET ?", (limit, skip))
    rows = cursor.fetchall()
    return jsonify([dict(row) for row in rows])

@app.route("/api/achievements/<int:achievement_id>", methods=["GET", "PUT", "DELETE"])
def achievement_detail(achievement_id):
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT * FROM achievements WHERE id = ?", (achievement_id,))
    achievement = cursor.fetchone()
    if achievement is None:
        return jsonify({"error": "Achievement not found"}), 404

    if request.method == "GET":
        return jsonify(dict(achievement))

    if request.method == "PUT":
        data = request.get_json() or {}
        description = data.get("description", achievement["description"]).strip()
        category = data.get("category", achievement["category"]).strip()
        tags = data.get("tags", achievement["tags"])
        tags_csv = ",".join(tags) if isinstance(tags, list) else str(tags)
        cursor.execute(
            "UPDATE achievements SET description = ?, category = ?, tags = ? WHERE id = ?",
            (description, category, tags_csv, achievement_id)
        )
        db.commit()
        cursor.execute("SELECT * FROM achievements WHERE id = ?", (achievement_id,))
        updated = cursor.fetchone()
        return jsonify(dict(updated))

    cursor.execute("DELETE FROM achievements WHERE id = ?", (achievement_id,))
    db.commit()
    return jsonify({"message": "Achievement deleted successfully"})

@app.route("/api/tasks", methods=["GET", "POST"])
def tasks():
    if request.method == "POST":
        data = request.get_json() or {}
        description = data.get("description", "").strip()
        priority = data.get("priority", "medium")
        status = data.get("status", "todo")
        due_date = data.get("due_date")

        if not description:
            return jsonify({"error": "description is required"}), 400

        created_date = datetime.utcnow().isoformat()
        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            "INSERT INTO tasks (description, priority, status, due_date, created_date) VALUES (?, ?, ?, ?, ?)",
            (description, priority, status, due_date, created_date)
        )
        db.commit()
        task_id = cursor.lastrowid
        cursor.execute("SELECT * FROM tasks WHERE id = ?", (task_id,))
        task = cursor.fetchone()
        return jsonify(dict(task)), 201

    skip = int(request.args.get("skip", 0))
    limit = int(request.args.get("limit", 100))
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT * FROM tasks ORDER BY created_date DESC LIMIT ? OFFSET ?", (limit, skip))
    rows = cursor.fetchall()
    return jsonify([dict(row) for row in rows])

@app.route("/api/tasks/<int:task_id>", methods=["GET", "PUT", "DELETE"])
def task_detail(task_id):
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT * FROM tasks WHERE id = ?", (task_id,))
    task = cursor.fetchone()
    if task is None:
        return jsonify({"error": "Task not found"}), 404

    if request.method == "GET":
        return jsonify(dict(task))

    if request.method == "PUT":
        data = request.get_json() or {}
        description = data.get("description", task["description"]).strip()
        priority = data.get("priority", task["priority"])
        status = data.get("status", task["status"])
        due_date = data.get("due_date", task["due_date"])
        cursor.execute(
            "UPDATE tasks SET description = ?, priority = ?, status = ?, due_date = ? WHERE id = ?",
            (description, priority, status, due_date, task_id)
        )
        db.commit()
        cursor.execute("SELECT * FROM tasks WHERE id = ?", (task_id,))
        updated = cursor.fetchone()
        return jsonify(dict(updated))

    cursor.execute("DELETE FROM tasks WHERE id = ?", (task_id,))
    db.commit()
    return jsonify({"message": "Task deleted successfully"})

@app.route("/api/export/achievements", methods=["POST"])
def export_achievements():
    data = request.get_json() or {}
    start_date = data.get("start_date")
    end_date = data.get("end_date")

    db = get_db()
    cursor = db.cursor()
    query = "SELECT * FROM achievements"
    params = []
    if start_date or end_date:
        conditions = []
        if start_date:
            conditions.append("timestamp >= ?")
            params.append(start_date)
        if end_date:
            conditions.append("timestamp <= ?")
            params.append(end_date)
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY timestamp DESC"
    cursor.execute(query, tuple(params))
    achievements = cursor.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Achievements"
    headers = ["ID", "Description", "Category", "Timestamp", "Tags"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for row_num, achievement in enumerate(achievements, 2):
        ws.cell(row=row_num, column=1, value=achievement["id"])
        ws.cell(row=row_num, column=2, value=achievement["description"])
        ws.cell(row=row_num, column=3, value=achievement["category"])
        ws.cell(row=row_num, column=4, value=achievement["timestamp"])
        ws.cell(row=row_num, column=5, value=achievement["tags"])

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value is not None and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[column_letter].width = max_length + 2

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()

    @after_this_request
    def remove_file(response):
        try:
            os.remove(temp_file.name)
        except OSError:
            pass
        return response

    filename = f"achievements_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(temp_file.name, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route("/api/export/tasks", methods=["POST"])
def export_tasks():
    data = request.get_json() or {}
    start_date = data.get("start_date")
    end_date = data.get("end_date")

    db = get_db()
    cursor = db.cursor()
    query = "SELECT * FROM tasks"
    params = []

    if start_date or end_date:
        conditions = []
        if start_date:
            conditions.append("created_date >= ?")
            params.append(start_date)
        if end_date:
            conditions.append("created_date <= ?")
            params.append(end_date)
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY created_date DESC"
    cursor.execute(query, tuple(params))
    tasks = cursor.fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tasks"
    headers = ["ID", "Description", "Priority", "Status", "Due Date", "Created Date"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    for row_num, task in enumerate(tasks, 2):
        ws.cell(row=row_num, column=1, value=task["id"])
        ws.cell(row=row_num, column=2, value=task["description"])
        ws.cell(row=row_num, column=3, value=task["priority"])
        ws.cell(row=row_num, column=4, value=task["status"])
        ws.cell(row=row_num, column=5, value=task["due_date"])
        ws.cell(row=row_num, column=6, value=task["created_date"])
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value is not None and len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        ws.column_dimensions[column_letter].width = max_length + 2

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(temp_file.name)
    temp_file.close()

    @after_this_request
    def remove_file(response):
        try:
            os.remove(temp_file.name)
        except OSError:
            pass
        return response

    filename = f"tasks_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(temp_file.name, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument-spreadsheetml.sheet')
